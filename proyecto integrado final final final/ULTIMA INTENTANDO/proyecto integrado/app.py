from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, send_file, Response
import sqlite3
import os
import hashlib
import re
import io
import csv
from functools import wraps
from datetime import datetime, timedelta

# Importar werkzeug para hash seguro de contraseñas
from werkzeug.security import generate_password_hash, check_password_hash

# MySQL support para producción
try:
    import pymysql
    pymysql.install_as_MySQLdb()
    MYSQL_AVAILABLE = True
except ImportError:
    MYSQL_AVAILABLE = False

# Constantes de configuración
IVA_CHILE = 0.19
MAX_RESULTS_PER_PAGE = 500

# Detectar entorno
ENVIRONMENT = os.environ.get('ENVIRONMENT', 'development')
USE_MYSQL = ENVIRONMENT == 'production' and MYSQL_AVAILABLE

app = Flask(__name__)

# Configuración de seguridad - usar variable de entorno en producción
app.secret_key = os.environ.get('SECRET_KEY', 'teknetau_dev_key_cambiar_en_produccion_2025')
app.config['DATABASE'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database', 'teknetau.db')
app.config['SESSION_COOKIE_SECURE'] = ENVIRONMENT == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Configuración MySQL para producción
MYSQL_CONFIG = {
    'host': os.environ.get('MYSQL_HOST', 'localhost'),
    'user': os.environ.get('MYSQL_USER', 'root'),
    'password': os.environ.get('MYSQL_PASSWORD', ''),
    'database': os.environ.get('MYSQL_DATABASE', 'teknetau_db'),
    'charset': 'utf8mb4',
    'use_unicode': True,
}

# Crear directorios necesarios
os.makedirs(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database'), exist_ok=True)
os.makedirs(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'css'), exist_ok=True)
os.makedirs(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'js'), exist_ok=True)
os.makedirs(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads'), exist_ok=True)


def get_db_connection():
    """Obtener conexión a base de datos (MySQL en producción, SQLite en desarrollo)."""
    if USE_MYSQL:
        try:
            import MySQLdb
            conn = MySQLdb.connect(**MYSQL_CONFIG)
            conn.set_charset('utf8mb4')
            return conn
        except Exception as e:
            # Fallback a SQLite si falla MySQL
            print(f"⚠️ Error conectando a MySQL: {e}. Usando SQLite...")
            return sqlite3.connect(app.config['DATABASE'])
    else:
        return sqlite3.connect(app.config['DATABASE'])


def rows_to_dicts(cursor, rows):
    """Convertir filas de la BD a diccionarios."""
    columns = [description[0] for description in cursor.description]
    return [dict(zip(columns, row)) for row in rows]


def normalize_rut(rut_str):
    """Normalizar RUT: quitar puntos y guiones."""
    if not rut_str:
        return ''
    return re.sub(r'[.-]', '', rut_str).strip()


def ensure_column(cursor, table, column_name, column_definition):
    """Agregar columna si no existe (migración simple)."""
    cursor.execute(f"PRAGMA table_info({table})")
    existing = {row[1] for row in cursor.fetchall()}
    if column_name not in existing:
        cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column_name} {column_definition}")

def init_database():
    """Inicializar base de datos según diagrama ER completo"""
    conn = sqlite3.connect(app.config['DATABASE'])
    cursor = conn.cursor()
    
    # ============================================================
    # TABLAS DE REFERENCIA GEOGRÁFICA
    # ============================================================
    
    # Tabla REGION
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS region (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            codigo TEXT UNIQUE
        )
    ''')
    
    # Tabla CIUDAD
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ciudad (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            region_id INTEGER,
            FOREIGN KEY (region_id) REFERENCES region(id)
        )
    ''')
    
    # Tabla COMUNA
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS comuna (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            ciudad_id INTEGER,
            FOREIGN KEY (ciudad_id) REFERENCES ciudad(id)
        )
    ''')
    
    # Tabla DIRECCION
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS direccion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            calle TEXT,
            numero TEXT,
            depto TEXT,
            codigo_postal TEXT,
            comuna_id INTEGER,
            FOREIGN KEY (comuna_id) REFERENCES comuna(id)
        )
    ''')
    
    # ============================================================
    # TABLAS DE TIPOS Y REFERENCIAS
    # ============================================================
    
    # Tabla TIPO_DOCUMENTO
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tipo_documento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            afecto_iva BOOLEAN DEFAULT 1
        )
    ''')
    
    # Tabla FORMA_PAGO
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS forma_pago (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            dias_plazo INTEGER DEFAULT 0
        )
    ''')
    
    # Tabla TIPO_PAGO
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tipo_pago (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            descripcion TEXT
        )
    ''')
    
    # Tabla ESTADO_DOCUMENTO
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS estado_documento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            color TEXT DEFAULT '#6c757d'
        )
    ''')
    
    # ============================================================
    # TABLA EMPRESA_PERSONA (CLIENTES)
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS empresa_persona (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rut TEXT UNIQUE NOT NULL,
            razon_social TEXT NOT NULL,
            nombre_fantasia TEXT,
            giro TEXT,
            telefono TEXT,
            email TEXT,
            sitio_web TEXT,
            direccion_id INTEGER,
            es_empresa BOOLEAN DEFAULT 1,
            fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP,
            activo BOOLEAN DEFAULT 1,
            observaciones TEXT,
            FOREIGN KEY (direccion_id) REFERENCES direccion(id)
        )
    ''')
    
    # ============================================================
    # TABLA USUARIOS (para login)
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            nombre TEXT,
            email TEXT,
            rol TEXT DEFAULT 'admin',
            activo BOOLEAN DEFAULT 1,
            fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # ============================================================
    # TABLA PROYECTOS
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS proyectos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            cliente_rut TEXT,
            empresa_persona_id INTEGER,
            presupuesto REAL DEFAULT 0,
            fecha_inicio DATE,
            fecha_termino DATE,
            estado TEXT DEFAULT 'Activo',
            observaciones TEXT,
            FOREIGN KEY (empresa_persona_id) REFERENCES empresa_persona(id)
        )
    ''')
    
    # ============================================================
    # TABLA PRODUCTOS_SERVICIOS
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS producto_servicio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            precio_unitario REAL DEFAULT 0,
            unidad_medida TEXT DEFAULT 'UN',
            es_servicio BOOLEAN DEFAULT 1,
            activo BOOLEAN DEFAULT 1,
            fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # ============================================================
    # TABLA DOCUMENTO (FACTURA, BOLETA, NC, ND)
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_doc INTEGER NOT NULL,
            tipo_documento_id INTEGER,
            fecha_emision DATE NOT NULL,
            fecha_vencimiento DATE,
            empresa_persona_id INTEGER,
            proyecto_id INTEGER,
            forma_pago_id INTEGER,
            estado_id INTEGER,
            subtotal REAL DEFAULT 0,
            descuento_porcentaje REAL DEFAULT 0,
            descuento_monto REAL DEFAULT 0,
            valor_neto REAL DEFAULT 0,
            iva REAL DEFAULT 0,
            valor_total REAL DEFAULT 0,
            observaciones TEXT,
            documento_referencia_id INTEGER,
            fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (tipo_documento_id) REFERENCES tipo_documento(id),
            FOREIGN KEY (empresa_persona_id) REFERENCES empresa_persona(id),
            FOREIGN KEY (proyecto_id) REFERENCES proyectos(id),
            FOREIGN KEY (forma_pago_id) REFERENCES forma_pago(id),
            FOREIGN KEY (estado_id) REFERENCES estado_documento(id),
            FOREIGN KEY (documento_referencia_id) REFERENCES documento(id)
        )
    ''')
    
    # ============================================================
    # TABLA DETALLE_DOCUMENTO (Líneas de cada documento)
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS detalle_documento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            documento_id INTEGER NOT NULL,
            producto_servicio_id INTEGER,
            descripcion TEXT,
            cantidad REAL DEFAULT 1,
            precio_unitario REAL DEFAULT 0,
            descuento_porcentaje REAL DEFAULT 0,
            descuento_monto REAL DEFAULT 0,
            subtotal REAL DEFAULT 0,
            orden INTEGER DEFAULT 0,
            FOREIGN KEY (documento_id) REFERENCES documento(id) ON DELETE CASCADE,
            FOREIGN KEY (producto_servicio_id) REFERENCES producto_servicio(id)
        )
    ''')
    
    # ============================================================
    # TABLA TRANSACCIONES (Pagos y movimientos)
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transaccion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            documento_id INTEGER,
            tipo_pago_id INTEGER,
            fecha_transaccion DATE NOT NULL,
            monto REAL NOT NULL,
            numero_referencia TEXT,
            banco TEXT,
            observaciones TEXT,
            fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (documento_id) REFERENCES documento(id),
            FOREIGN KEY (tipo_pago_id) REFERENCES tipo_pago(id)
        )
    ''')
    
    # ============================================================
    # TABLA TRANSPORTE (para guías de despacho)
    # ============================================================
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transporte (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            documento_id INTEGER,
            patente TEXT,
            chofer_nombre TEXT,
            chofer_rut TEXT,
            direccion_destino_id INTEGER,
            fecha_despacho DATE,
            observaciones TEXT,
            FOREIGN KEY (documento_id) REFERENCES documento(id),
            FOREIGN KEY (direccion_destino_id) REFERENCES direccion(id)
        )
    ''')
    
    # ============================================================
    # MIGRACIÓN: Mantener compatibilidad con tablas antiguas
    # ============================================================
    
    # Verificar si existen las tablas antiguas y migrar datos
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='clientes'")
    if cursor.fetchone():
        # Migrar clientes a empresa_persona
        cursor.execute('''
            INSERT OR IGNORE INTO empresa_persona (rut, razon_social, giro, telefono, email, activo)
            SELECT rut, razon_social, giro, telefono, email, activo FROM clientes
        ''')
    
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='documentos'")
    if cursor.fetchone():
        # Los documentos antiguos se mantienen en la tabla 'documentos' por compatibilidad
        # Se pueden migrar manualmente después
        pass
    
    # ============================================================
    # DATOS INICIALES - Regiones de Chile
    # ============================================================
    
    regiones_chile = [
        ('XV', 'Arica y Parinacota'),
        ('I', 'Tarapacá'),
        ('II', 'Antofagasta'),
        ('III', 'Atacama'),
        ('IV', 'Coquimbo'),
        ('V', 'Valparaíso'),
        ('RM', 'Metropolitana de Santiago'),
        ('VI', "O'Higgins"),
        ('VII', 'Maule'),
        ('XVI', 'Ñuble'),
        ('VIII', 'Biobío'),
        ('IX', 'La Araucanía'),
        ('XIV', 'Los Ríos'),
        ('X', 'Los Lagos'),
        ('XI', 'Aysén'),
        ('XII', 'Magallanes')
    ]
    
    for codigo, nombre in regiones_chile:
        cursor.execute('INSERT OR IGNORE INTO region (codigo, nombre) VALUES (?, ?)', (codigo, nombre))
    
    # Ciudades principales (Región Metropolitana como ejemplo)
    cursor.execute('SELECT id FROM region WHERE codigo = ?', ('RM',))
    rm = cursor.fetchone()
    if rm:
        ciudades_rm = ['Santiago', 'Puente Alto', 'Maipú', 'Las Condes', 'La Florida', 'Providencia', 'Vitacura']
        for ciudad in ciudades_rm:
            cursor.execute('INSERT OR IGNORE INTO ciudad (nombre, region_id) VALUES (?, ?)', (ciudad, rm[0]))
    
    # Comunas de Santiago
    cursor.execute('SELECT id FROM ciudad WHERE nombre = ?', ('Santiago',))
    stgo = cursor.fetchone()
    if stgo:
        comunas_stgo = ['Santiago Centro', 'Providencia', 'Las Condes', 'Vitacura', 'Ñuñoa', 'La Reina', 
                        'Macul', 'Peñalolén', 'La Florida', 'Puente Alto', 'San Miguel', 'San Joaquín',
                        'Maipú', 'Cerrillos', 'Estación Central', 'Quinta Normal', 'Renca', 'Quilicura']
        for comuna in comunas_stgo:
            cursor.execute('INSERT OR IGNORE INTO comuna (nombre, ciudad_id) VALUES (?, ?)', (comuna, stgo[0]))
    
    # ============================================================
    # DATOS INICIALES - Tipos de documento
    # ============================================================
    
    tipos_doc = [
        ('FAC', 'Factura Electrónica', 'Factura afecta a IVA', 1),
        ('FEX', 'Factura Exenta', 'Factura exenta de IVA', 0),
        ('BOL', 'Boleta Electrónica', 'Boleta de venta', 1),
        ('NC', 'Nota de Crédito', 'Nota de crédito electrónica', 1),
        ('ND', 'Nota de Débito', 'Nota de débito electrónica', 1),
        ('GD', 'Guía de Despacho', 'Guía de despacho electrónica', 0)
    ]
    
    for codigo, nombre, desc, iva in tipos_doc:
        cursor.execute('INSERT OR IGNORE INTO tipo_documento (codigo, nombre, descripcion, afecto_iva) VALUES (?, ?, ?, ?)',
                      (codigo, nombre, desc, iva))
    
    # ============================================================
    # DATOS INICIALES - Formas de pago
    # ============================================================
    
    formas_pago = [
        ('CONT', 'Contado', 'Pago al contado', 0),
        ('30D', 'Crédito 30 días', 'Pago a 30 días', 30),
        ('60D', 'Crédito 60 días', 'Pago a 60 días', 60),
        ('90D', 'Crédito 90 días', 'Pago a 90 días', 90)
    ]
    
    for codigo, nombre, desc, dias in formas_pago:
        cursor.execute('INSERT OR IGNORE INTO forma_pago (codigo, nombre, descripcion, dias_plazo) VALUES (?, ?, ?, ?)',
                      (codigo, nombre, desc, dias))
    
    # ============================================================
    # DATOS INICIALES - Tipos de pago
    # ============================================================
    
    tipos_pago = [
        ('EFE', 'Efectivo', 'Pago en efectivo'),
        ('TRF', 'Transferencia', 'Transferencia bancaria'),
        ('CHQ', 'Cheque', 'Pago con cheque'),
        ('TDC', 'Tarjeta de Crédito', 'Pago con tarjeta de crédito'),
        ('TDD', 'Tarjeta de Débito', 'Pago con tarjeta de débito'),
        ('CRE', 'Crédito', 'Pago a crédito')
    ]
    
    for codigo, nombre, desc in tipos_pago:
        cursor.execute('INSERT OR IGNORE INTO tipo_pago (codigo, nombre, descripcion) VALUES (?, ?, ?)',
                      (codigo, nombre, desc))
    
    # ============================================================
    # DATOS INICIALES - Estados de documento
    # ============================================================
    
    estados = [
        ('PEN', 'Pendiente', 'Documento pendiente de pago', '#ffc107'),
        ('PAG', 'Pagado', 'Documento pagado completamente', '#28a745'),
        ('PAR', 'Pago Parcial', 'Documento con pago parcial', '#17a2b8'),
        ('ANU', 'Anulado', 'Documento anulado', '#dc3545'),
        ('VEN', 'Vencido', 'Documento vencido sin pagar', '#fd7e14')
    ]
    
    for codigo, nombre, desc, color in estados:
        cursor.execute('INSERT OR IGNORE INTO estado_documento (codigo, nombre, descripcion, color) VALUES (?, ?, ?, ?)',
                      (codigo, nombre, desc, color))
    
    # ============================================================
    # USUARIO POR DEFECTO
    # ============================================================
    
    # Usar werkzeug para hash seguro de contraseñas
    password_hash = generate_password_hash('admin123', method='pbkdf2:sha256')
    cursor.execute('INSERT OR IGNORE INTO usuarios (username, password_hash, nombre, rol) VALUES (?, ?, ?, ?)', 
                  ('admin', password_hash, 'Administrador', 'admin'))
    
    # ============================================================
    # DATOS DE EJEMPLO - Empresa/Persona
    # ============================================================
    
    cursor.execute('''INSERT OR IGNORE INTO empresa_persona (rut, razon_social, giro, email, es_empresa, activo) 
                     VALUES ('76660180-4', 'WINPY SPA', 'Tecnología', 'contacto.winpy@gmail.com', 1, 1)''')
    cursor.execute('''INSERT OR IGNORE INTO empresa_persona (rut, razon_social, giro, email, es_empresa, activo) 
                     VALUES ('78138410-0', 'APLICACIONES COMPUTACIONALES SPA', 'Software', 'contacto.aplicaciones@gmail.com', 1, 1)''')
    
    # ============================================================
    # COMPATIBILIDAD - Mantener tablas antiguas
    # ============================================================
    
    # Tabla clientes (compatibilidad hacia atrás)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rut TEXT UNIQUE,
            razon_social TEXT,
            giro TEXT,
            telefono TEXT,
            email TEXT,
            direccion TEXT,
            comuna TEXT,
            fecha_creacion TEXT DEFAULT CURRENT_TIMESTAMP,
            activo BOOLEAN DEFAULT 1
        )
    ''')
    
    # Tabla documentos antigua (compatibilidad hacia atrás)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_doc INTEGER,
            tipo_doc TEXT,
            fecha_emision DATE,
            cliente_rut TEXT,
            descripcion TEXT,
            valor_neto REAL,
            iva REAL,
            valor_total REAL,
            estado TEXT DEFAULT 'Pendiente',
            forma_pago TEXT,
            proyecto_codigo TEXT
        )
    ''')
    
    # Asegurar columnas in tablas existentes
    ensure_column(cursor, 'clientes', 'telefono', 'TEXT')
    ensure_column(cursor, 'clientes', 'email', 'TEXT')
    ensure_column(cursor, 'clientes', 'direccion', 'TEXT')
    ensure_column(cursor, 'clientes', 'comuna', 'TEXT')
    ensure_column(cursor, 'clientes', 'fecha_creacion', 'TEXT DEFAULT CURRENT_TIMESTAMP')
    ensure_column(cursor, 'clientes', 'cuenta_corriente', 'TEXT')
    ensure_column(cursor, 'clientes', 'idt', 'TEXT')
    ensure_column(cursor, 'clientes', 'desc_idt', 'TEXT')
    ensure_column(cursor, 'documentos', 'proyecto_codigo', 'TEXT')
    ensure_column(cursor, 'proyectos', 'cliente_rut', 'TEXT')
    ensure_column(cursor, 'proyectos', 'empresa_persona_id', 'INTEGER')
    ensure_column(cursor, 'clientes', 'banco', 'TEXT')
    ensure_column(cursor, 'clientes', 'observaciones', 'TEXT')
    
    # Sincronizar clientes con empresa_persona
    cursor.execute('''
        INSERT OR IGNORE INTO clientes (rut, razon_social, giro, telefono, email, activo)
        SELECT rut, razon_social, giro, telefono, email, activo FROM empresa_persona
    ''')
    
    conn.commit()
    conn.close()

init_database()


def get_db_connection():
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    return conn


def rows_to_dicts(rows):
    resultado = []
    for row in rows:
        if isinstance(row, sqlite3.Row):
            resultado.append({key: row[key] for key in row.keys()})
        elif isinstance(row, dict):
            resultado.append(row)
        else:
            try:
                resultado.append(dict(row))
            except Exception:
                resultado.append({'value': row})
    return resultado


def normalize_rut(rut: str) -> str:
    if not rut:
        return ''
    clean = rut.upper().replace('.', '').replace('-', '').replace('\u00A0', '').replace(' ', '')
    if len(clean) < 2:
        return ''
    cuerpo, dv = clean[:-1], clean[-1]
    return f"{int(cuerpo)}-{dv}" if cuerpo.isdigit() else f"{cuerpo}-{dv}"


def validate_rut(rut: str) -> bool:
    """Valida RUT chileno usando el algoritmo correcto del dígito verificador"""
    if not rut:
        return False
    
    # Limpiar el RUT: eliminar puntos, guiones y espacios
    clean = rut.upper().replace('.', '').replace('-', '').replace(' ', '').strip()
    
    if len(clean) < 2:
        return False
    
    # Separar cuerpo y dígito verificador
    cuerpo = clean[:-1]
    dv = clean[-1]
    
    # Validar que el cuerpo sea solo números
    if not cuerpo.isdigit():
        return False
    
    # Algoritmo de validación del dígito verificador
    suma = 0
    multiplicador = 2
    
    # Recorrer de derecha a izquierda con multiplicadores 2,3,4,5,6,7,2,3...
    for i in range(len(cuerpo) - 1, -1, -1):
        suma += int(cuerpo[i]) * multiplicador
        multiplicador += 1
        if multiplicador > 7:
            multiplicador = 2
    
    # Calcular dígito verificador: 11 - (suma % 11)
    resto = suma % 11
    dv_calculado = 11 - resto
    
    # Casos especiales
    if dv_calculado == 11:
        dv_esperado = '0'
    elif dv_calculado == 10:
        dv_esperado = 'K'
    else:
        dv_esperado = str(dv_calculado)
    
    return dv_esperado == dv


def validate_email(email: str) -> bool:
    """Valida correo electrónico - acepta gmail.com, .cl y otros dominios comunes"""
    if not email:
        return False
    email = email.strip()
    # Acepta cualquier correo válido con dominios comunes
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.(com|cl|net|org|edu|gov|io|co)$'
    return bool(re.match(pattern, email, re.IGNORECASE))


def validate_telefono_chileno(telefono: str) -> bool:
    """Valida teléfono celular chileno: +56 9 XXXX XXXX o 9 XXXX XXXX"""
    if not telefono:
        return True  # Teléfono es opcional
    telefono = telefono.strip().replace(' ', '').replace('-', '')
    # Formato: +56912345678 o 912345678
    pattern = r'^(\+?56)?9[0-9]{8}$'
    return bool(re.match(pattern, telefono))


def format_telefono_chileno(telefono: str) -> str:
    """Formatea teléfono chileno a formato estándar +56 9 XXXX XXXX"""
    if not telefono:
        return ''
    telefono = telefono.strip().replace(' ', '').replace('-', '').replace('+', '')
    if telefono.startswith('56'):
        telefono = telefono[2:]
    if len(telefono) == 9 and telefono.startswith('9'):
        return f"+56 {telefono[0]} {telefono[1:5]} {telefono[5:]}"
    return telefono


# Decorador para login requerido
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            # Para APIs, devolver JSON en lugar de redirect
            if request.path.startswith('/api/'):
                return jsonify({'error': 'No autorizado', 'login_required': True}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(required_roles):
    """Decorador para verificar el rol del usuario.
    
    Uso:
        @role_required('admin')  # Solo admin
        @role_required(['admin', 'usuario'])  # Admin o usuario
    """
    if isinstance(required_roles, str):
        required_roles = [required_roles]
    
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Primero verificar que esté logueado
            if 'user_id' not in session:
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'No autorizado', 'login_required': True}), 401
                return redirect(url_for('login'))
            
            # Luego verificar el rol
            user_role = session.get('rol', 'usuario')
            if user_role not in required_roles:
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Acceso denegado. Rol requerido: ' + ', '.join(required_roles)}), 403
                flash('No tienes permiso para acceder a esta sección', 'danger')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# RUTAS PRINCIPALES
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        conn = sqlite3.connect(app.config['DATABASE'])
        cursor = conn.cursor()
        cursor.execute('SELECT id, username, password_hash, nombre, rol FROM usuarios WHERE username = ?', (username,))
        user = cursor.fetchone()
        conn.close()
        
        # Verificar contraseña - soporta hash antiguo SHA256 y nuevo werkzeug
        password_valid = False
        if user:
            stored_hash = user[2]
            # Intentar con werkzeug primero (formato pbkdf2:sha256)
            if stored_hash.startswith('pbkdf2:') or stored_hash.startswith('scrypt:'):
                password_valid = check_password_hash(stored_hash, password)
            else:
                # Fallback a SHA256 para usuarios antiguos
                import hashlib
                password_valid = stored_hash == hashlib.sha256(password.encode()).hexdigest()
        
        if user and password_valid:
            session['user_id'] = user[0]
            session['username'] = user[1]
            session['nombre'] = user[3]
            session['rol'] = user[4] if len(user) > 4 else 'usuario'
            flash('¡Bienvenido ' + user[3] + '!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    session.clear()
    flash('Sesión cerrada exitosamente', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    conn = get_db_connection()
    
    stats = {
        'total_clientes': conn.execute("SELECT COUNT(*) FROM clientes WHERE activo = 1").fetchone()[0],
        'total_proyectos': conn.execute("SELECT COUNT(*) FROM proyectos WHERE estado = 'Activo'").fetchone()[0],
        'documentos_pendientes': conn.execute("SELECT COUNT(*) FROM documentos WHERE estado = 'Pendiente'").fetchone()[0],
        'deuda_total': conn.execute("SELECT COALESCE(SUM(valor_total), 0) FROM documentos WHERE estado = 'Pendiente'").fetchone()[0] or 0
    }
    
    documentos_recientes = rows_to_dicts(conn.execute('''
        SELECT d.*, c.razon_social FROM documentos d 
        LEFT JOIN clientes c ON d.cliente_rut = c.rut 
        ORDER BY d.fecha_emision DESC LIMIT 5
    ''').fetchall())
    
    clientes = rows_to_dicts(conn.execute("SELECT rut, razon_social FROM clientes WHERE activo = 1").fetchall())
    conn.close()
    
    return render_template('index.html', stats=stats, documentos=documentos_recientes, clientes=clientes)

@app.route('/clientes')
@login_required
def clientes():
    # Lista de bancos de Chile
    bancos_chile = [
        'Banco de Chile',
        'BCI',
        'Scotiabank',
        'BBVA',
        'Santander',
        'Itaú',
        'Corpbanc',
        'Banco del Estado',
        'Banco Bice',
        'Banco Security',
        'Banco Falabella',
        'Banco Ripley',
        'Banco Consigo',
        'Banco Inbursa',
        'Banco Azteca',
        'Banco Caja Social',
        'Banco Desarrollo',
        'Banco Finterra',
        'Banco Multiva',
        'Banco PagaTodo',
        'Hsbc',
        'Deutsche Bank',
        'Rabobank',
        'Banco Crédito e Inversiones',
        'Otro'
    ]
    
    conn = get_db_connection()
    clientes = rows_to_dicts(conn.execute("SELECT * FROM clientes WHERE activo = 1").fetchall())
    conn.close()
    return render_template('clientes.html', clientes=clientes, bancos=sorted(bancos_chile))

@app.route('/administracion')
@role_required('admin')
def administracion():
    """Página de administración: cambiar contraseña y gestionar perfiles"""
    return render_template('administracion.html')

@app.route('/facturas')
@login_required
def facturas():
    conn = get_db_connection()
    clientes = rows_to_dicts(conn.execute("SELECT rut, razon_social FROM clientes WHERE activo = 1").fetchall())
    proyectos = rows_to_dicts(conn.execute("SELECT codigo, nombre, cliente_rut FROM proyectos WHERE estado = 'Activo'").fetchall())
    
    ultima_factura = conn.execute("SELECT COALESCE(MAX(numero_doc), 0) FROM documentos WHERE tipo_doc = 'FAC'").fetchone()[0]
    conn.close()
    
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('facturas.html', clientes=clientes, proyectos=proyectos, proximo_numero=ultima_factura + 1, today=today)

@app.route('/boletas')
@login_required
def boletas():
    conn = get_db_connection()
    clientes = rows_to_dicts(conn.execute("SELECT rut, razon_social FROM clientes WHERE activo = 1").fetchall())
    proyectos = rows_to_dicts(conn.execute("SELECT codigo, nombre, cliente_rut FROM proyectos WHERE estado = 'Activo'").fetchall())
    
    ultima_boleta = conn.execute("SELECT COALESCE(MAX(numero_doc), 0) FROM documentos WHERE tipo_doc = 'BOL'").fetchone()[0]
    conn.close()
    
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('boletas.html', clientes=clientes, proyectos=proyectos, proximo_numero=ultima_boleta + 1, today=today)


@app.route('/proyectos')
@login_required
def proyectos():
    conn = get_db_connection()
    # Mostrar todos los proyectos con información del cliente vinculado
    proyectos_list = rows_to_dicts(conn.execute('''
        SELECT p.*, c.razon_social
        FROM proyectos p
        LEFT JOIN clientes c ON c.rut = p.cliente_rut
        ORDER BY p.nombre
    ''').fetchall())
    clientes_list = rows_to_dicts(conn.execute('SELECT rut, razon_social FROM clientes WHERE activo = 1 ORDER BY razon_social').fetchall())
    conn.close()
    return render_template('proyectos.html', proyectos=proyectos_list, clientes=clientes_list)


@app.route('/notas-credito')
@login_required
def notas_credito():
    conn = get_db_connection()
    clientes_list = rows_to_dicts(conn.execute("SELECT rut, razon_social FROM clientes WHERE activo = 1").fetchall())
    conn.close()
    return render_template('notas_credito.html', clientes=clientes_list)


@app.route('/notas-debito')
@login_required
def notas_debito():
    conn = get_db_connection()
    clientes_list = rows_to_dicts(conn.execute("SELECT rut, razon_social FROM clientes WHERE activo = 1").fetchall())
    conn.close()
    return render_template('notas_debito.html', clientes=clientes_list)


@app.route('/reportes')
@login_required
def reportes():
    conn = get_db_connection()
    cursor = conn.cursor()
    # Cambiar consulta para incluir todos los proyectos, no solo los activos
    cursor.execute('SELECT codigo, nombre, estado FROM proyectos ORDER BY nombre')
    proyectos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template('reportes.html', proyectos=proyectos)


@app.route('/reportes-completos')
@login_required
def reportes_completos():
    # Redirigir a la página de reportes unificada
    return redirect(url_for('reportes'))


@app.route('/cambiar-password', methods=['GET', 'POST'])
@role_required('admin')
def cambiar_password():
    if request.method == 'POST':
        password_actual = request.form.get('password_actual')
        nuevo_password = request.form.get('nuevo_password')
        confirmar_password = request.form.get('confirmar_password')

        if nuevo_password != confirmar_password:
            flash('La nueva contraseña no coincide con la confirmación.', 'error')
            return redirect(url_for('cambiar_password'))

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT password_hash FROM usuarios WHERE id = ?', (session['user_id'],))
        row = cursor.fetchone()
        password_hash_actual = row['password_hash'] if row else None

        # Verificar contraseña actual - soporta ambos formatos
        password_valid = False
        if password_hash_actual:
            if password_hash_actual.startswith('pbkdf2:') or password_hash_actual.startswith('scrypt:'):
                password_valid = check_password_hash(password_hash_actual, password_actual)
            else:
                import hashlib
                password_valid = password_hash_actual == hashlib.sha256(password_actual.encode()).hexdigest()
        
        if not password_valid:
            conn.close()
            flash('La contraseña actual es incorrecta.', 'error')
            return redirect(url_for('cambiar_password'))

        # Guardar nueva contraseña con hash seguro
        nuevo_hash = generate_password_hash(nuevo_password, method='pbkdf2:sha256')
        cursor.execute('UPDATE usuarios SET password_hash = ? WHERE id = ?', (nuevo_hash, session['user_id']))
        conn.commit()
        conn.close()

        flash('Contraseña actualizada correctamente.', 'success')
        return redirect(url_for('index'))

    return render_template('cambiar_password.html')


# APIS
@app.route('/api/generar-documento', methods=['POST'])
@login_required
def api_generar_documento():
    data = request.get_json()
    conn = sqlite3.connect(app.config['DATABASE'])
    
    try:
        cursor = conn.cursor()
        
        # Obtener proyecto_codigo si se especificó
        proyecto_codigo = data.get('proyecto_codigo', '') or None
        
        cursor.execute('''
            INSERT INTO documentos (numero_doc, tipo_doc, fecha_emision, cliente_rut, descripcion, 
                                 valor_neto, iva, valor_total, estado, forma_pago, proyecto_codigo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Pendiente', ?, ?)
        ''', (
            data['numero_doc'],
            data['tipo_doc'],
            data['fecha_emision'],
            data['cliente_rut'],
            data['descripcion'],
            data['valor_neto'],
            data['iva'],
            data['valor_total'],
            data.get('forma_pago', 'Contado'),
            proyecto_codigo
        ))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Documento generado exitosamente'})
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})


# API DOCUMENTOS PENDIENTES (para dashboard)
@app.route('/api/documentos-pendientes')
@login_required
def api_documentos_pendientes():
    """Obtiene los documentos pendientes de pago para el dashboard"""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT d.id, d.numero_doc, d.tipo_doc, d.fecha_emision, d.valor_total, 
                   d.estado, c.razon_social
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE d.estado = 'Pendiente'
            ORDER BY d.fecha_emision DESC
            LIMIT 10
        ''')
        rows = cursor.fetchall()
        documentos = []
        for row in rows:
            documentos.append({
                'id': row[0],
                'numero_doc': row[1],
                'tipo_doc': row[2],
                'fecha_emision': row[3],
                'valor_total': row[4],
                'estado': row[5],
                'cliente': row[6] or 'Sin cliente'
            })
        return jsonify(documentos)
    finally:
        conn.close()


# REPORTES APIs - SIMPLIFICADOS Y OPTIMIZADOS
@app.route('/api/reporte-deudas')
@login_required
def api_reporte_deudas():
    try:
        conn = sqlite3.connect(app.config['DATABASE'])
        deudas = conn.execute('''
            SELECT c.rut, c.razon_social, COUNT(d.id) as cantidad, COALESCE(SUM(d.valor_total), 0) as total
            FROM clientes c
            LEFT JOIN documentos d ON c.rut = d.cliente_rut AND d.estado = 'Pendiente'
            WHERE c.activo = 1
            GROUP BY c.rut, c.razon_social
            HAVING total > 0
        ''').fetchall()
        conn.close()
        
        resultado = []
        for deuda in deudas:
            resultado.append({
                'rut': deuda[0],
                'razon_social': deuda[1],
                'cantidad_documentos': deuda[2],
                'total_deuda': float(deuda[3])
            })
        
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reporte-ventas-mensual')
@login_required
def api_reporte_ventas_mensual():
    try:
        conn = sqlite3.connect(app.config['DATABASE'])
        ventas = conn.execute('''
            SELECT strftime('%Y-%m', fecha_emision) as mes, tipo_doc, 
                   COUNT(*) as cantidad, SUM(valor_total) as total
            FROM documentos 
            WHERE tipo_doc IN ('FAC', 'BOL') AND estado != 'Anulado'
            GROUP BY mes, tipo_doc
            ORDER BY mes DESC
        ''').fetchall()
        conn.close()
        
        resultado = []
        for venta in ventas:
            resultado.append({
                'mes': venta[0],
                'tipo_doc': venta[1],
                'cantidad': venta[2],
                'total_ventas': float(venta[3] or 0)
            })
        
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reporte-top-clientes')
@login_required
def api_reporte_top_clientes():
    try:
        conn = sqlite3.connect(app.config['DATABASE'])
        clientes = conn.execute('''
            SELECT c.razon_social, c.rut, COUNT(d.id) as documentos, SUM(d.valor_total) as total
            FROM clientes c
            LEFT JOIN documentos d ON c.rut = d.cliente_rut 
            WHERE d.estado != 'Anulado' AND d.tipo_doc IN ('FAC', 'BOL')
            GROUP BY c.rut, c.razon_social
            ORDER BY total DESC
            LIMIT 10
        ''').fetchall()
        conn.close()
        
        resultado = []
        for cliente in clientes:
            resultado.append({
                'razon_social': cliente[0],
                'rut': cliente[1],
                'total_documentos': cliente[2],
                'total_facturado': float(cliente[3] or 0)
            })
        
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reporte-anual')
@login_required
def api_reporte_anual():
    try:
        conn = sqlite3.connect(app.config['DATABASE'])
        
        # Ventas anuales
        ventas = conn.execute('''
            SELECT strftime('%Y', fecha_emision) as año, strftime('%m', fecha_emision) as mes,
                   COUNT(*) as cantidad, SUM(valor_total) as total
            FROM documentos 
            WHERE tipo_doc IN ('FAC', 'BOL') AND estado != 'Anulado'
            GROUP BY año, mes
        ''').fetchall()
        
        # Resumen anual
        resumen = conn.execute('''
            SELECT strftime('%Y', fecha_emision) as año, COUNT(*) as documentos,
                   SUM(valor_total) as ventas, SUM(CASE WHEN estado = 'Pagado' THEN valor_total ELSE 0 END) as pagado
            FROM documentos 
            WHERE tipo_doc IN ('FAC', 'BOL') AND estado != 'Anulado'
            GROUP BY año
        ''').fetchall()
        
        conn.close()
        
        reporte_anual = {
            'ventas_anuales': [{'año': v[0], 'mes': v[1], 'cantidad_documentos': v[2], 'total_ventas': float(v[3] or 0)} for v in ventas],
            'resumen_anual': [{'año': r[0], 'total_documentos': r[1], 'total_ventas': float(r[2] or 0), 'total_pagado': float(r[3] or 0)} for r in resumen]
        };
        
        return jsonify(reporte_anual)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/clientes', methods=['GET', 'POST', 'DELETE'])
@app.route('/api-dev/clientes', methods=['GET', 'POST', 'DELETE'])
@login_required
def api_clientes():
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        if request.method == 'GET':
            estado = request.args.get('estado', '').lower()
            if estado == 'activo':
                query = "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo FROM clientes WHERE activo = 1 ORDER BY razon_social"
                rows = cursor.execute(query).fetchall()
            elif estado == 'inactivo':
                query = "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo FROM clientes WHERE activo = 0 ORDER BY razon_social"
                rows = cursor.execute(query).fetchall()
            else:
                query = "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo FROM clientes ORDER BY razon_social"
                rows = cursor.execute(query).fetchall()
            return jsonify(rows_to_dicts(rows))

        if request.method == 'DELETE':
            rut = request.args.get('rut')
            if not rut:
                return jsonify({'success': False, 'error': 'RUT requerido'}), 400
            rut_norm = normalize_rut(rut)
            if USE_MYSQL:
                cursor.execute('UPDATE clientes SET activo = 0 WHERE rut = %s', (rut_norm,))
            else:
                cursor.execute('UPDATE clientes SET activo = 0 WHERE rut = ?', (rut_norm,))
            conn.commit()
            return jsonify({'success': cursor.rowcount > 0})

        # POST - Crear nuevo cliente
        data = request.get_json() or {}
        rut = (data.get('rut') or '').strip()
        razon_social = (data.get('razon_social') or '').strip()
        email = (data.get('email') or '').strip()

        # Validaciones básicas
        if not rut or not razon_social:
            return jsonify({'success': False, 'error': 'RUT y razón social son obligatorios'}), 400

        # Validar RUT con módulo 11
        if not validate_rut(rut):
            return jsonify({'success': False, 'error': 'RUT inválido. Formato: XX.XXX.XXX-X o similar'}), 400

        if not email or not validate_email(email):
            return jsonify({'success': False, 'error': 'Debe ingresar un correo válido (.com, .cl, etc.)'}), 400

        telefono = (data.get('telefono') or '').strip()
        if telefono and not validate_telefono_chileno(telefono):
            return jsonify({'success': False, 'error': 'Teléfono inválido. Use formato celular chileno: +56 9 XXXX XXXX'}), 400

        # Normalizar datos
        rut_norm = normalize_rut(rut)
        email_norm = email.lower()
        telefono_fmt = format_telefono_chileno(telefono)
        giro = (data.get('giro') or '').strip()
        direccion = (data.get('direccion') or '').strip()
        comuna = (data.get('comuna') or '').strip()
        banco = (data.get('banco') or '').strip()
        cuenta_corriente = (data.get('cuenta_corriente') or '').strip()
        observaciones = (data.get('observaciones') or '').strip()

        # Buscar cliente existente
        try:
            if USE_MYSQL:
                cursor.execute(
                    "SELECT id FROM clientes WHERE rut = %s",
                    (rut_norm,)
                )
                existente = cursor.fetchone()
            else:
                existente = cursor.execute(
                    "SELECT id FROM clientes WHERE rut = ?",
                    (rut_norm,)
                ).fetchone()

            if existente:
                # Actualizar cliente existente
                if USE_MYSQL:
                    cursor.execute(
                        '''
                        UPDATE clientes
                        SET razon_social = %s, giro = %s, telefono = %s, email = %s, 
                            direccion = %s, comuna = %s, cuenta_corriente = %s, banco = %s, observaciones = %s, activo = 1
                        WHERE rut = %s
                        ''',
                        (razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones, rut_norm)
                    )
                else:
                    cursor.execute(
                        '''
                        UPDATE clientes
                        SET razon_social = ?, giro = ?, telefono = ?, email = ?, 
                            direccion = ?, comuna = ?, cuenta_corriente = ?, banco = ?, observaciones = ?, activo = 1
                        WHERE rut = ?
                        ''',
                        (razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones, rut_norm)
                    )
                mensaje = 'Cliente actualizado correctamente'
            else:
                # Insertar nuevo cliente
                if USE_MYSQL:
                    cursor.execute(
                        '''
                        INSERT INTO clientes (rut, razon_social, giro, telefono, email, 
                                            direccion, comuna, cuenta_corriente, banco, observaciones, activo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
                        ''',
                        (rut_norm, razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones)
                    )
                else:
                    cursor.execute(
                        '''
                        INSERT INTO clientes (rut, razon_social, giro, telefono, email, 
                                            direccion, comuna, cuenta_corriente, banco, observaciones, activo)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                        ''',
                        (rut_norm, razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones)
                    )
                mensaje = 'Cliente creado correctamente'

            conn.commit()
            return jsonify({'success': True, 'message': mensaje})

        except Exception as e:
            conn.rollback()
            error_msg = str(e)
            print(f"Error al guardar cliente: {error_msg}")  # Para debugging
            
            if 'UNIQUE constraint failed' in error_msg:
                return jsonify({'success': False, 'error': 'El RUT ya existe en el sistema'}), 400
            return jsonify({'success': False, 'error': f'Error al guardar: {error_msg}'}), 400

    except Exception as e:
        print(f"Error general en api_clientes: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/clientes-dev', methods=['GET', 'POST', 'DELETE'])
def api_clientes_dev():
    """Versión de desarrollo del API sin autenticación requerida"""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        if request.method == 'GET':
            estado = request.args.get('estado', '').lower()
            if estado == 'activo':
                query = "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo FROM clientes WHERE activo = 1 ORDER BY razon_social"
                rows = cursor.execute(query).fetchall()
            elif estado == 'inactivo':
                query = "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo FROM clientes WHERE activo = 0 ORDER BY razon_social"
                rows = cursor.execute(query).fetchall()
            else:
                query = "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo FROM clientes ORDER BY razon_social"
                rows = cursor.execute(query).fetchall()
            return jsonify(rows_to_dicts(rows))

        if request.method == 'DELETE':
            rut = request.args.get('rut')
            if not rut:
                return jsonify({'success': False, 'error': 'RUT requerido'}), 400
            rut_norm = normalize_rut(rut)
            if USE_MYSQL:
                cursor.execute('UPDATE clientes SET activo = 0 WHERE rut = %s', (rut_norm,))
            else:
                cursor.execute('UPDATE clientes SET activo = 0 WHERE rut = ?', (rut_norm,))
            conn.commit()
            return jsonify({'success': cursor.rowcount > 0})

        # POST - Crear nuevo cliente
        data = request.get_json() or {}
        rut = (data.get('rut') or '').strip()
        razon_social = (data.get('razon_social') or '').strip()
        email = (data.get('email') or '').strip()

        # Validaciones básicas
        if not rut or not razon_social:
            return jsonify({'success': False, 'error': 'RUT y razón social son obligatorios'}), 400

        # Validar RUT con módulo 11
        if not validate_rut(rut):
            return jsonify({'success': False, 'error': 'RUT inválido. Formato: XX.XXX.XXX-X o similar'}), 400

        if not email or not validate_email(email):
            return jsonify({'success': False, 'error': 'Debe ingresar un correo válido (.com, .cl, etc.)'}), 400

        telefono = (data.get('telefono') or '').strip()
        if telefono and not validate_telefono_chileno(telefono):
            return jsonify({'success': False, 'error': 'Teléfono inválido. Use formato celular chileno: +56 9 XXXX XXXX'}), 400

        # Normalizar datos
        rut_norm = normalize_rut(rut)
        email_norm = email.lower()
        telefono_fmt = format_telefono_chileno(telefono)
        giro = (data.get('giro') or '').strip()
        direccion = (data.get('direccion') or '').strip()
        comuna = (data.get('comuna') or '').strip()
        banco = (data.get('banco') or '').strip()
        cuenta_corriente = (data.get('cuenta_corriente') or '').strip()
        observaciones = (data.get('observaciones') or '').strip()
        activo = data.get('activo', 1)  # Por defecto activo = 1 si no se especifica

        # Buscar cliente existente
        if USE_MYSQL:
            cursor.execute(
                "SELECT id FROM clientes WHERE rut = %s",
                (rut_norm,)
            )
            existente = cursor.fetchone()
        else:
            existente = cursor.execute(
                "SELECT id FROM clientes WHERE rut = ?",
                (rut_norm,)
            ).fetchone()

        try:
            if existente:
                # Actualizar cliente existente
                if USE_MYSQL:
                    cursor.execute(
                        '''
                        UPDATE clientes
                        SET razon_social = %s, giro = %s, telefono = %s, email = %s, 
                            direccion = %s, comuna = %s, cuenta_corriente = %s, banco = %s, observaciones = %s, activo = %s
                        WHERE rut = %s
                        ''',
                        (razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones, activo, rut_norm)
                    )
                else:
                    cursor.execute(
                        '''
                        UPDATE clientes
                        SET razon_social = ?, giro = ?, telefono = ?, email = ?, 
                            direccion = ?, comuna = ?, cuenta_corriente = ?, banco = ?, observaciones = ?, activo = ?
                        WHERE rut = ?
                        ''',
                        (razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones, activo, rut_norm)
                    )
                mensaje = 'Cliente actualizado correctamente'
            else:
                # Insertar nuevo cliente
                if USE_MYSQL:
                    cursor.execute(
                        '''
                        INSERT INTO clientes (rut, razon_social, giro, telefono, email, 
                                            direccion, comuna, cuenta_corriente, banco, observaciones, activo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
                        ''',
                        (rut_norm, razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones)
                    )
                else:
                    cursor.execute(
                        '''
                        INSERT INTO clientes (rut, razon_social, giro, telefono, email, 
                                            direccion, comuna, cuenta_corriente, banco, observaciones, activo)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                        ''',
                        (rut_norm, razon_social, giro, telefono_fmt, email_norm, direccion, comuna, cuenta_corriente, banco, observaciones)
                    )
                mensaje = 'Cliente creado correctamente'

            conn.commit()
            return jsonify({'success': True, 'message': mensaje})

        except Exception as e:
            conn.rollback()
            error_msg = str(e)
            print(f"Error al guardar cliente: {error_msg}")  # Para debugging
            
            if 'UNIQUE constraint failed' in error_msg:
                return jsonify({'success': False, 'error': 'El RUT ya existe en el sistema'}), 400
            return jsonify({'success': False, 'error': f'Error al guardar: {error_msg}'}), 400

    except Exception as e:
        print(f"Error general en api_clientes_dev: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/clientes-dev/<rut>', methods=['GET', 'PUT'])
def api_clientes_dev_get_update(rut):
    """Versión de desarrollo del API sin autenticación requerida"""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        rut_norm = normalize_rut(rut)
        
        if request.method == 'GET':
            # Buscar cliente por RUT normalizado
            row = cursor.execute(
                "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo "
                "FROM clientes WHERE rut = ? AND activo = 1",
                (rut_norm,)
            ).fetchone()
            
            if not row:
                return jsonify({'error': 'Cliente no encontrado'}), 404
            
            # Convertir a dict
            resultado = rows_to_dicts([row])[0]
            return jsonify(resultado)
            
        elif request.method == 'PUT':
            data = request.get_json() or {}
            razon_social = (data.get('razon_social') or '').strip()
            email = (data.get('email') or '').strip()
            telefono = (data.get('telefono') or '').strip()
            giro = (data.get('giro') or '').strip()
            direccion = (data.get('direccion') or '').strip()
            comuna = (data.get('comuna') or '').strip()
            cuenta_corriente = (data.get('cuenta_corriente') or '').strip()
            banco = (data.get('banco') or '').strip()
            observaciones = (data.get('observaciones') or '').strip()

            # Validaciones
            if not razon_social:
                return jsonify({'success': False, 'error': 'La razón social es obligatoria'}), 400
            
            if not validate_rut(rut_norm):
                return jsonify({'success': False, 'error': 'RUT inválido'}), 400
            
            if not email or not validate_email(email):
                return jsonify({'success': False, 'error': 'Debe ingresar un correo válido (.com, .cl, etc.)'}), 400
            
            if telefono and not validate_telefono_chileno(telefono):
                return jsonify({'success': False, 'error': 'Teléfono inválido. Use formato celular chileno: +56 9 XXXX XXXX'}), 400

            # Actualizar cliente por RUT normalizado
            cursor.execute(
                '''
                UPDATE clientes
                SET razon_social = ?, giro = ?, telefono = ?, email = ?, 
                    direccion = ?, comuna = ?, cuenta_corriente = ?, banco = ?, observaciones = ?, activo = 1
                WHERE rut = ?
                ''',
                (
                    razon_social,
                    giro,
                    format_telefono_chileno(telefono),
                    email.lower(),
                    direccion,
                    comuna,
                    cuenta_corriente,
                    banco,
                    observaciones,
                    rut_norm
                )
            )
            conn.commit()
            
            if cursor.rowcount > 0:
                return jsonify({'success': True})
            else:
                return jsonify({'success': False, 'error': 'Cliente no encontrado'}), 404
                
    except Exception as e:
        print(f"Error en api_clientes_dev_get_update: {str(e)}")
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500
    finally:
        conn.close()


@app.route('/api/clientes/<rut>', methods=['GET', 'PUT'])
@login_required
def api_clientes_get_update(rut):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        rut_norm = normalize_rut(rut)
        
        if request.method == 'GET':
            # Buscar cliente por RUT normalizado
            row = cursor.execute(
                "SELECT id, rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente, banco, observaciones, activo "
                "FROM clientes WHERE rut = ? AND activo = 1",
                (rut_norm,)
            ).fetchone()
            
            if not row:
                return jsonify({'error': 'Cliente no encontrado'}), 404
            
            # Convertir a dict
            resultado = rows_to_dicts([row])[0]
            return jsonify(resultado)
            
        elif request.method == 'PUT':
            data = request.get_json() or {}
            razon_social = (data.get('razon_social') or '').strip()
            email = (data.get('email') or '').strip()
            telefono = (data.get('telefono') or '').strip()
            giro = (data.get('giro') or '').strip()
            direccion = (data.get('direccion') or '').strip()
            comuna = (data.get('comuna') or '').strip()
            cuenta_corriente = (data.get('cuenta_corriente') or '').strip()
            banco = (data.get('banco') or '').strip()
            observaciones = (data.get('observaciones') or '').strip()

            # Validaciones
            if not razon_social:
                return jsonify({'success': False, 'error': 'La razón social es obligatoria'}), 400
            
            if not validate_rut(rut_norm):
                return jsonify({'success': False, 'error': 'RUT inválido'}), 400
            
            if not email or not validate_email(email):
                return jsonify({'success': False, 'error': 'Debe ingresar un correo válido (.com, .cl, etc.)'}), 400
            
            if telefono and not validate_telefono_chileno(telefono):
                return jsonify({'success': False, 'error': 'Teléfono inválido. Use formato celular chileno: +56 9 XXXX XXXX'}), 400

            # Actualizar cliente por RUT normalizado
            cursor.execute(
                '''
                UPDATE clientes
                SET razon_social = ?, giro = ?, telefono = ?, email = ?, 
                    direccion = ?, comuna = ?, cuenta_corriente = ?, banco = ?, observaciones = ?, activo = 1
                WHERE rut = ?
                ''',
                (
                    razon_social,
                    giro,
                    format_telefono_chileno(telefono),
                    email.lower(),
                    direccion,
                    comuna,
                    cuenta_corriente,
                    banco,
                    observaciones,
                    rut_norm
                )
            )
            conn.commit()
            
            if cursor.rowcount > 0:
                return jsonify({'success': True})
            else:
                return jsonify({'success': False, 'error': 'Cliente no encontrado'}), 404
                
    except Exception as e:
        print(f"Error en api_clientes_get_update: {str(e)}")
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500
    finally:
        conn.close()


# ============================================================
# APIS DE ADMINISTRACIÓN (USUARIOS Y CONTRASEÑAS)
# ============================================================

@app.route('/api/cambiar-password', methods=['POST'])
@role_required('admin')
def cambiar_password_api():
    """Cambiar contraseña del usuario actual"""
    data = request.get_json() or {}
    password_actual = data.get('password_actual', '').strip()
    password_nueva = data.get('password_nueva', '').strip()
    
    if not password_actual or not password_nueva:
        return jsonify({'success': False, 'error': 'Contraseñas requeridas'}), 400
    
    if len(password_nueva) < 8:
        return jsonify({'success': False, 'error': 'Contraseña debe tener mínimo 8 caracteres'}), 400
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener usuario actual
        usuario_id = session.get('user_id')
        usuario = cursor.execute(
            "SELECT password_hash FROM usuarios WHERE id = ?",
            (usuario_id,)
        ).fetchone()
        
        if not usuario:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        
        # Verificar contraseña actual
        if not check_password_hash(usuario['password_hash'], password_actual):
            return jsonify({'success': False, 'error': 'Contraseña actual incorrecta'}), 401
        
        # Actualizar contraseña
        nueva_hash = generate_password_hash(password_nueva)
        cursor.execute(
            "UPDATE usuarios SET password_hash = ? WHERE id = ?",
            (nueva_hash, usuario_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Contraseña actualizada correctamente'})
        
    except Exception as e:
        print(f"Error al cambiar password: {str(e)}")
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


@app.route('/api/usuarios-dev', methods=['GET', 'POST', 'DELETE'])
@role_required('admin')
def api_usuarios_dev():
    """API de usuarios para desarrollo - gestionar perfiles (solo admin)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if request.method == 'GET':
            # Listar todos los usuarios
            rows = cursor.execute(
                "SELECT id, username, nombre, email, rol, activo, fecha_creacion FROM usuarios ORDER BY username"
            ).fetchall()
            conn.close()
            return jsonify(rows_to_dicts(rows))
        
        if request.method == 'DELETE':
            # Eliminar usuario
            usuario_id = request.args.get('id')
            if not usuario_id:
                return jsonify({'success': False, 'error': 'ID requerido'}), 400
            
            # No permitir eliminar el usuario actual
            if int(usuario_id) == session.get('user_id'):
                return jsonify({'success': False, 'error': 'No puedes eliminar tu propio usuario'}), 403
            
            cursor.execute('DELETE FROM usuarios WHERE id = ?', (usuario_id,))
            conn.commit()
            conn.close()
            
            return jsonify({'success': cursor.rowcount > 0})
        
        # POST - Crear o actualizar usuario
        data = request.get_json() or {}
        username = (data.get('username') or '').strip()
        nombre = (data.get('nombre') or '').strip()
        email = (data.get('email') or '').strip()
        rol = (data.get('rol') or 'usuario').strip()
        password = (data.get('password') or '').strip()
        activo = data.get('activo', 1)
        usuario_id = data.get('id')
        
        # Validaciones
        if not username or not nombre or not email:
            return jsonify({'success': False, 'error': 'Username, nombre y email son obligatorios'}), 400
        
        if not validate_email(email):
            return jsonify({'success': False, 'error': 'Email inválido'}), 400
        
        if usuario_id:
            # Actualizar usuario existente
            if password and len(password) < 8:
                return jsonify({'success': False, 'error': 'Contraseña debe tener mínimo 8 caracteres'}), 400
            
            if password:
                password_hash = generate_password_hash(password)
                cursor.execute(
                    '''
                    UPDATE usuarios
                    SET nombre = ?, email = ?, rol = ?, activo = ?, password_hash = ?
                    WHERE id = ?
                    ''',
                    (nombre, email, rol, activo, password_hash, usuario_id)
                )
            else:
                cursor.execute(
                    '''
                    UPDATE usuarios
                    SET nombre = ?, email = ?, rol = ?, activo = ?
                    WHERE id = ?
                    ''',
                    (nombre, email, rol, activo, usuario_id)
                )
            mensaje = 'Usuario actualizado correctamente'
        else:
            # Crear nuevo usuario
            if not password or len(password) < 8:
                return jsonify({'success': False, 'error': 'Contraseña requerida (mínimo 8 caracteres)'}), 400
            
            password_hash = generate_password_hash(password)
            
            # Verificar que username sea único
            existente = cursor.execute(
                "SELECT id FROM usuarios WHERE username = ?",
                (username,)
            ).fetchone()
            
            if existente:
                return jsonify({'success': False, 'error': 'El usuario ya existe'}), 400
            
            cursor.execute(
                '''
                INSERT INTO usuarios (username, nombre, email, rol, password_hash, activo)
                VALUES (?, ?, ?, ?, ?, ?)
                ''',
                (username, nombre, email, rol, password_hash, activo)
            )
            mensaje = 'Usuario creado correctamente'
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': mensaje})
        
    except Exception as e:
        print(f"Error en api_usuarios_dev: {str(e)}")
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


@app.route('/api/proyectos', methods=['GET', 'POST'])
@login_required
def api_proyectos():
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        if request.method == 'GET':
            rows = cursor.execute('''
                SELECT p.*, c.razon_social
                FROM proyectos p
                LEFT JOIN clientes c ON c.rut = p.cliente_rut
                ORDER BY COALESCE(p.fecha_inicio, p.id) DESC
            ''').fetchall()
            return jsonify(rows_to_dicts(rows))

        data = request.get_json() or {}
        codigo = (data.get('codigo') or '').strip()
        nombre = (data.get('nombre') or '').strip()
        cliente_rut = (data.get('cliente_rut') or '').strip()
        fecha_inicio = (data.get('fecha_inicio') or '').strip()

        if not codigo or not nombre or not cliente_rut or not fecha_inicio:
            return jsonify({'success': False, 'error': 'Código, nombre, cliente y fecha de inicio son obligatorios'}), 400

        cliente_rut_norm = normalize_rut(cliente_rut)
        if not validate_rut(cliente_rut_norm):
            return jsonify({'success': False, 'error': f'El RUT del cliente es inválido: {cliente_rut}'}), 400

        # Verificar que el cliente existe - ser más flexible en la búsqueda
        existe_cliente = cursor.execute(
            "SELECT id, rut, razon_social FROM clientes WHERE rut = ?",
            (cliente_rut_norm,)
        ).fetchone()
        
        if not existe_cliente:
            # Si no existe en clientes, buscar en empresa_persona
            existe_cliente = cursor.execute(
                "SELECT id, rut FROM empresa_persona WHERE rut = ?",
                (cliente_rut_norm,)
            ).fetchone()
            
            if existe_cliente:
                # Insertar en clientes desde empresa_persona
                cursor.execute('''
                    INSERT OR IGNORE INTO clientes (rut, razon_social, giro, telefono, email, activo)
                    SELECT rut, razon_social, giro, email, activo FROM empresa_persona WHERE rut = ?
                ''', (cliente_rut_norm,))
                conn.commit()
            else:
                return jsonify({'success': False, 'error': f'El cliente con RUT {cliente_rut_norm} no existe en el sistema'}), 404

        try:
            presupuesto = float(data.get('presupuesto') or 0)
        except ValueError:
            return jsonify({'success': False, 'error': 'El presupuesto debe ser numérico'}), 400

        fecha_termino = data.get('fecha_termino') or None
        estado = (data.get('estado') or 'Activo').strip() or 'Activo'

        existente = cursor.execute('SELECT id FROM proyectos WHERE codigo = ?', (codigo,)).fetchone()

        if existente:
            cursor.execute('''
                UPDATE proyectos
                SET nombre = ?, descripcion = ?, cliente_rut = ?, presupuesto = ?, fecha_inicio = ?, fecha_termino = ?, estado = ?
                WHERE codigo = ?
            ''', (
                nombre,
                data.get('descripcion'),
                cliente_rut_norm,
                presupuesto,
                fecha_inicio,
                fecha_termino,
                estado,
                codigo
            ))
        else:
            cursor.execute('''
                INSERT INTO proyectos (codigo, nombre, descripcion, cliente_rut, presupuesto, fecha_inicio, fecha_termino, estado)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                codigo,
                nombre,
                data.get('descripcion'),
                cliente_rut_norm,
                presupuesto,
                fecha_inicio,
                fecha_termino,
                estado
            ))

        conn.commit()
        return jsonify({'success': True})

    finally:
        conn.close()


@app.route('/api/proyectos/<codigo>', methods=['PUT', 'DELETE'])
@login_required
def api_proyectos_update_delete(codigo):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if request.method == 'DELETE':
            cursor.execute('DELETE FROM proyectos WHERE codigo = ?', (codigo,))
            conn.commit()
            if cursor.rowcount > 0:
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': 'Proyecto no encontrado'}), 404
        
        # PUT - Actualizar proyecto
        data = request.get_json() or {}
        nombre = (data.get('nombre') or '').strip()
        cliente_rut = (data.get('cliente_rut') or '').strip()
        fecha_inicio = (data.get('fecha_inicio') or '').strip()

        if not nombre or not cliente_rut or not fecha_inicio:
            return jsonify({'success': False, 'error': 'Nombre, cliente y fecha de inicio son obligatorios'}), 400

        cliente_rut_norm = normalize_rut(cliente_rut)
        if not validate_rut(cliente_rut_norm):
            return jsonify({'success': False, 'error': 'El RUT del cliente es inválido'}), 400

        try:
            presupuesto = float(data.get('presupuesto') or 0)
        except ValueError:
            return jsonify({'success': False, 'error': 'El presupuesto debe ser numérico'}), 400

        fecha_termino = data.get('fecha_termino') or None
        estado = (data.get('estado') or 'Activo').strip() or 'Activo'

        cursor.execute('''
            UPDATE proyectos
            SET nombre = ?, descripcion = ?, cliente_rut = ?, presupuesto = ?, fecha_inicio = ?, fecha_termino = ?, estado = ?
            WHERE codigo = ?
        ''', (
            nombre,
            data.get('descripcion'),
            cliente_rut_norm,
            presupuesto,
            fecha_inicio,
            fecha_termino,
            estado,
            codigo
        ))
        conn.commit()
        
        if cursor.rowcount > 0:
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Proyecto no encontrado'}), 404

    finally:
        conn.close()


@app.route('/api/proyecto/<codigo>/progreso')
@login_required
def api_proyecto_progreso(codigo):
    """Obtiene el progreso de pago de un proyecto"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener datos del proyecto
        proyecto = cursor.execute('''
            SELECT p.*, c.razon_social 
            FROM proyectos p
            LEFT JOIN clientes c ON p.cliente_rut = c.rut
            WHERE p.codigo = ?
        ''', (codigo,)).fetchone()
        
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        
        presupuesto = proyecto['presupuesto'] or 0
        
        # Obtener documentos asociados al proyecto
        documentos = cursor.execute('''
            SELECT d.*, c.razon_social as cliente_nombre
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE d.proyecto_codigo = ?
            ORDER BY d.fecha_emision DESC
        ''', (codigo,)).fetchall()
        
        # =====================================================
        # LÓGICA SIMPLIFICADA Y CORRECTA
        # =====================================================
        
        # 1. Total facturado = suma de FAC/BOL que NO están anulados
        total_facturado = sum(
            doc['valor_total'] or 0 
            for doc in documentos 
            if doc['tipo_doc'] in ('FAC', 'BOL') and doc['estado'] != 'Anulado'
        )
        
        # 2. Total pagado = suma de FAC/BOL con estado 'Pagado'
        total_pagado = sum(
            doc['valor_total'] or 0 
            for doc in documentos 
            if doc['tipo_doc'] in ('FAC', 'BOL') and doc['estado'] == 'Pagado'
        )
        
        # 3. Total pendiente = facturado - pagado (solo lo que falta por pagar de lo facturado)
        total_pendiente = max(0, total_facturado - total_pagado)
        
        # 4. Calcular porcentaje de pago
        #    - Si hay presupuesto: % = pagado / presupuesto * 100
        #    - Si no hay presupuesto pero hay facturación: % = pagado / facturado * 100
        #    - Si pagado > presupuesto: 100% (excedieron el presupuesto)
        if presupuesto > 0:
            if total_pagado >= presupuesto:
                porcentaje_pagado = 100  # Pagaron todo el presupuesto o más
            else:
                porcentaje_pagado = (total_pagado / presupuesto) * 100
        elif total_facturado > 0:
            porcentaje_pagado = (total_pagado / total_facturado) * 100
        else:
            porcentaje_pagado = 0
        
        # Limitar a 100%
        porcentaje_pagado = min(100, porcentaje_pagado)
        
        return jsonify({
            'proyecto': {
                'codigo': proyecto['codigo'],
                'nombre': proyecto['nombre'],
                'cliente': proyecto['razon_social'],
                'presupuesto': presupuesto,
                'estado': proyecto['estado']
            },
            'progreso': {
                'total_facturado': total_facturado,
                'total_pagado': total_pagado,
                'total_pendiente': total_pendiente,
                'porcentaje_pagado': round(porcentaje_pagado, 2),
                'porcentaje_restante': round(100 - porcentaje_pagado, 2)
            },
            'documentos': [dict(doc) for doc in documentos],
            'cantidad_documentos': len(documentos)
        })
    finally:
        conn.close()


@app.route('/api/proyecto/<codigo>/documentos')
@login_required
def api_proyecto_documentos(codigo):
    """Obtiene todos los documentos asociados a un proyecto"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        documentos = cursor.execute('''
            SELECT d.*, c.razon_social as cliente_nombre
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE d.proyecto_codigo = ?
            ORDER BY d.fecha_emision DESC
        ''', (codigo,)).fetchall()
        
        return jsonify([dict(doc) for doc in documentos])
    finally:
        conn.close()


@app.route('/api/documentos/<int:doc_id>/estado', methods=['PUT'])
@login_required
def api_documento_cambiar_estado(doc_id):
    """Cambia el estado de un documento (Pendiente/Pagado/Anulado/Aplicada)"""
    data = request.get_json() or {}
    nuevo_estado = data.get('estado', '').strip()
    
    if nuevo_estado not in ['Pendiente', 'Pagado', 'Anulado', 'Aplicada']:
        return jsonify({'success': False, 'error': 'Estado inválido'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('UPDATE documentos SET estado = ? WHERE id = ?', (nuevo_estado, doc_id))
        conn.commit()
        
        if cursor.rowcount > 0:
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
    finally:
        conn.close()


@app.route('/api/ultimos-documentos/<tipo_doc>')
@login_required
def api_ultimos_documentos(tipo_doc):
    """Obtiene los últimos documentos de un tipo específico"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT d.id, d.numero_doc, d.tipo_doc, d.fecha_emision, d.valor_total, 
                   d.estado, d.proyecto_codigo, c.razon_social as cliente_nombre
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE d.tipo_doc = ?
            ORDER BY d.fecha_emision DESC, d.id DESC
            LIMIT 20
        ''', (tipo_doc,))
        
        documentos = [dict(row) for row in cursor.fetchall()]
        return jsonify({'documentos': documentos})
    finally:
        conn.close()


@app.route('/api/buscar-documentos')
@login_required
def api_buscar_documentos():
    """Busca documentos con filtros"""
    tipo_doc = request.args.get('tipo_doc', '')
    estado = request.args.get('estado', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    proyecto = request.args.get('proyecto', '')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = '''
            SELECT d.id, d.numero_doc, d.tipo_doc, d.fecha_emision, d.cliente_rut,
                   d.descripcion, d.valor_neto, d.iva, d.valor_total, d.estado,
                   d.proyecto_codigo, c.razon_social
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE 1=1
        '''
        params = []
        
        if tipo_doc:
            query += ' AND d.tipo_doc = ?'
            params.append(tipo_doc)
        if estado:
            query += ' AND d.estado = ?'
            params.append(estado)
        if fecha_desde:
            query += ' AND d.fecha_emision >= ?'
            params.append(fecha_desde)
        if fecha_hasta:
            query += ' AND d.fecha_emision <= ?'
            params.append(fecha_hasta)
        if proyecto:
            query += ' AND d.proyecto_codigo = ?'
            params.append(proyecto)
        
        query += ' ORDER BY d.fecha_emision DESC LIMIT 500'
        
        cursor.execute(query, params)
        documentos = [dict(row) for row in cursor.fetchall()]
        return jsonify({'documentos': documentos})
    finally:
        conn.close()


@app.route('/api/top-clientes')
@login_required
def api_top_clientes():
    """Obtiene el top 10 de clientes por facturación"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT c.rut, c.razon_social,
                   COUNT(d.id) as total_documentos,
                   COALESCE(SUM(d.valor_total), 0) as total_facturado,
                   COALESCE(SUM(CASE WHEN d.estado = 'Pagado' THEN d.valor_total ELSE 0 END), 0) as total_pagado,
                   COALESCE(SUM(CASE WHEN d.estado = 'Pendiente' THEN d.valor_total ELSE 0 END), 0) as total_pendiente
            FROM clientes c
            LEFT JOIN documentos d ON c.rut = d.cliente_rut AND d.tipo_doc IN ('FAC', 'BOL')
            WHERE c.activo = 1
            GROUP BY c.rut, c.razon_social
            HAVING total_facturado > 0
            ORDER BY total_facturado DESC
            LIMIT 10
        ''').fetchall()
        conn.close()
        
        resultado = []
        for cliente in clientes:
            resultado.append({
                'razon_social': cliente[0],
                'rut': cliente[1],
                'total_documentos': cliente[2],
                'total_facturado': float(cliente[3] or 0),
                'total_pagado': float(cliente[4] or 0),
                'total_pendiente': float(cliente[5] or 0)
            })
        
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/exportar-deudas-excel')
@login_required
def api_exportar_deudas_excel():
    """Exporta reporte de deudas a Excel"""
    import io
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'Módulo openpyxl no instalado'}), 500
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Agregar cuenta_corriente del cliente
        cursor.execute('''
            SELECT d.numero_doc, d.tipo_doc, d.fecha_emision, c.razon_social, c.rut,
                   d.valor_total, d.proyecto_codigo, c.cuenta_corriente
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE d.estado = 'Pendiente' AND d.tipo_doc IN ('FAC', 'BOL', 'ND')
            ORDER BY c.razon_social, d.fecha_emision
        ''')
        rows = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deudas Pendientes"
        
        headers = ['N° Doc', 'Tipo', 'Fecha', 'Cliente', 'RUT', 'Monto', 'Proyecto', 'Cuenta Corriente']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reporte_deudas.xlsx'
        )
    finally:
        conn.close()


@app.route('/api/exportar-documentos-excel')
@login_required
def api_exportar_documentos_excel():
    """Exporta todos los documentos a Excel"""
    import io
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'Módulo openpyxl no instalado'}), 500
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Agregar cuenta_corriente del cliente
        cursor.execute('''
            SELECT d.numero_doc, d.tipo_doc, d.fecha_emision, c.razon_social,
                   d.valor_neto, d.iva, d.valor_total, d.estado, d.proyecto_codigo, c.cuenta_corriente
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            ORDER BY d.fecha_emision DESC
        ''')
        rows = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Documentos"
        
        headers = ['N° Doc', 'Tipo', 'Fecha', 'Cliente', 'Neto', 'IVA', 'Total', 'Estado', 'Proyecto', 'Cuenta Corriente']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='documentos.xlsx'
        )
    finally:
        conn.close()


# ==================== APIs de Exportación CSV ====================

@app.route('/api/cuentas-corrientes', methods=['GET'])
@login_required
def api_cuentas_corrientes():
    """Obtiene lista de cuentas corrientes únicas"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            'SELECT DISTINCT cuenta_corriente FROM clientes WHERE activo = 1 AND cuenta_corriente IS NOT NULL AND cuenta_corriente != "" ORDER BY cuenta_corriente'
        )
        rows = cursor.fetchall()
        cuentas = [row[0] for row in rows]
        return jsonify(cuentas)
    finally:
        conn.close()


@app.route('/api/cuenta-corriente-detalle/<cuenta>')
@login_required
def api_cuenta_corriente_detalle(cuenta):
    """Obtiene detalles de clientes, pagos y deudas para una cuenta corriente"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener clientes de la cuenta corriente
        cursor.execute(
            'SELECT rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente FROM clientes WHERE activo = 1 AND cuenta_corriente = ? ORDER BY razon_social',
            (cuenta,)
        )
        
        rows = cursor.fetchall()
        columns = [description[0] for description in cursor.description]
        
        clientes = []
        ruts_clientes = []
        for row in rows:
            cliente = dict(zip(columns, row))
            clientes.append(cliente)
            ruts_clientes.append(cliente['rut'])
        
        # Obtener documentos (facturas, boletas, etc.) de los clientes de esta cuenta
        documentos = []
        pagos_totales = 0
        deudas_totales = 0
        
        if ruts_clientes:
            placeholders = ','.join('?' * len(ruts_clientes))
            cursor.execute(f'''
                SELECT d.*, c.razon_social
                FROM documentos d
                LEFT JOIN clientes c ON c.rut = d.cliente_rut
                WHERE d.cliente_rut IN ({placeholders})
                ORDER BY d.fecha_emision DESC
            ''', ruts_clientes)
            
            doc_rows = cursor.fetchall()
            doc_columns = [description[0] for description in cursor.description]
            
            for row in doc_rows:
                doc = dict(zip(doc_columns, row))
                documentos.append(doc)
                
                # Acumular totales
                valor_total = doc.get('valor_total') or 0
                if doc.get('estado') == 'Pagado':
                    pagos_totales += valor_total
                elif doc.get('estado') == 'Pendiente':
                    deudas_totales += valor_total
        
        return jsonify({
            'cuenta': cuenta,
            'clientes': clientes,
            'total_clientes': len(clientes),
            'documentos': documentos,
            'total_documentos': len(documentos),
            'total_pagos': pagos_totales,
            'total_deudas': deudas_totales,
            'total_facturado': pagos_totales + deudas_totales
        })
    finally:
        conn.close()


@app.route('/api/exportar-cuenta-corriente-csv/<cuenta>')
@login_required
def api_exportar_cuenta_corriente_csv(cuenta):
    """Exporta reporte completo de cuenta corriente a CSV con clientes, pagos y deudas"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener datos completos
        cursor.execute(
            'SELECT rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente FROM clientes WHERE activo = 1 AND cuenta_corriente = ? ORDER BY razon_social',
            (cuenta,)
        )
        
        clientes_rows = cursor.fetchall()
        ruts_clientes = [row[0] for row in clientes_rows]
        
        # Obtener documentos
        documentos = []
        if ruts_clientes:
            placeholders = ','.join('?' * len(ruts_clientes))
            cursor.execute(f'''
                SELECT d.numero_doc, d.tipo_doc, d.fecha_emision, d.cliente_rut, d.valor_neto, d.iva, d.valor_total, d.estado, c.razon_social
                FROM documentos d
                LEFT JOIN clientes c ON c.rut = d.cliente_rut
                WHERE d.cliente_rut IN ({placeholders})
                ORDER BY d.fecha_emision DESC
            ''', ruts_clientes)
            
            documentos = cursor.fetchall()
        
        # Crear CSV con codificación UTF-8 BOM para Excel
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator='\n')
        
        # Encabezado con resumen
        writer.writerow(['REPORTE DE CUENTA CORRIENTE'])
        writer.writerow(['Cuenta Corriente', cuenta])
        writer.writerow(['Fecha Generación', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow([''])
        
        # Sección de clientes
        writer.writerow(['CLIENTES ASOCIADOS A LA CUENTA'])
        writer.writerow(['RUT', 'Razón Social', 'Giro', 'Teléfono', 'Email', 'Dirección', 'Comuna'])
        for row in clientes_rows:
            writer.writerow(row)
        
        writer.writerow([''])
        writer.writerow(['DOCUMENTOS Y MOVIMIENTOS'])
        writer.writerow(['Número', 'Tipo', 'Fecha', 'Cliente RUT', 'Neto', 'IVA', 'Total', 'Estado', 'Razón Social'])
        
        pagos_totales = 0
        deudas_totales = 0
        
        for doc in documentos:
            doc_list = list(doc)
            if doc_list[4]:
                doc_list[4] = f"{float(doc_list[4]):,.2f}".replace(',', '.')
            if doc_list[5]:
                doc_list[5] = f"{float(doc_list[5]):,.2f}".replace(',', '.')
            if doc_list[6]:
                doc_list[6] = f"{float(doc_list[6]):,.2f}".replace(',', '.')
            writer.writerow(doc_list)
            valor_total = float(doc[6]) if doc[6] else 0
            if doc[7] == 'Pagado':
                pagos_totales += valor_total
            elif doc[7] == 'Pendiente':
                deudas_totales += valor_total
        
        writer.writerow([''])
        writer.writerow(['RESUMEN FINANCIERO'])
        writer.writerow(['Total Pagado', f"{pagos_totales:,.2f}".replace(',', '.')])
        writer.writerow(['Total Deuda', f"{deudas_totales:,.2f}".replace(',', '.')])
        writer.writerow(['Total Facturado', f"{(pagos_totales + deudas_totales):,.2f}".replace(',', '.')])
        
        # Agregar BOM UTF-8
        csv_content = '\ufeff' + output.getvalue()
        
        from flask import Response
        return Response(
            csv_content,
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename=cuenta_corriente_{cuenta}.csv'}
        )
    finally:
        conn.close()


@app.route('/api/exportar-clientes-csv')
@login_required
def api_exportar_clientes_csv():
    """Exporta clientes a CSV con filtro opcional por cuenta corriente"""
    # Obtener parámetro de filtro
    cuenta_corriente = request.args.get('cuenta_corriente', '').strip()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Construir query con filtro opcional
        if cuenta_corriente:
            # Filtrar por cuenta corriente exacta
            cursor.execute(
                'SELECT rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente FROM clientes WHERE activo = 1 AND cuenta_corriente = ? ORDER BY razon_social',
                (cuenta_corriente,)
            )
        else:
            # Todos los clientes
            cursor.execute('SELECT rut, razon_social, giro, telefono, email, direccion, comuna, cuenta_corriente FROM clientes WHERE activo = 1 ORDER BY razon_social')
        
        rows = cursor.fetchall()
        
        # Crear CSV con codificación UTF-8 BOM para Excel
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator='\n')
        
        # Encabezado
        if cuenta_corriente:
            titulo = f'Clientes - Cuenta Corriente {cuenta_corriente}'
        else:
            titulo = 'Listado de Clientes Activos'
        writer.writerow([titulo])
        writer.writerow(['Fecha', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow(['Total de registros', len(rows)])
        writer.writerow([''])
        
        # Headers
        writer.writerow(['RUT', 'Razón Social', 'Giro', 'Teléfono', 'Email', 'Dirección', 'Comuna', 'Cuenta Corriente'])
        
        for row in rows:
            writer.writerow(row)
        
        # Agregar BOM UTF-8
        csv_content = '\ufeff' + output.getvalue()
        
        # Nombrar el archivo según el filtro
        filename = 'clientes.csv'
        if cuenta_corriente:
            filename = f'clientes_{cuenta_corriente}.csv'
        
        from flask import Response
        return Response(
            csv_content,
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    finally:
        conn.close()


@app.route('/api/exportar-proyectos-csv')
@login_required
def api_exportar_proyectos_csv():
    """Exporta todos los proyectos a CSV"""
    import io
    import csv
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Agregar cuenta_corriente del cliente
        cursor.execute('''
            SELECT p.codigo, p.nombre, p.descripcion, c.razon_social, p.presupuesto, 
                   p.fecha_inicio, p.fecha_termino, p.estado, c.cuenta_corriente
            FROM proyectos p
            LEFT JOIN clientes c ON p.cliente_rut = c.rut
        ''')
        rows = cursor.fetchall()
        
        # Crear CSV con codificación UTF-8 BOM para Excel
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator='\n')
        
        # Encabezado
        writer.writerow(['Listado de Proyectos'])
        writer.writerow(['Fecha', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow(['Total de registros', len(rows)])
        writer.writerow([''])
        
        writer.writerow(['Código', 'Nombre', 'Descripción', 'Cliente', 'Presupuesto', 'Fecha Inicio', 'Fecha Término', 'Estado', 'Cuenta Corriente'])
        
        for row in rows:
            row_list = list(row)
            if row_list[4]:
                row_list[4] = f"{float(row_list[4]):,.2f}".replace(',', '.')
            writer.writerow(row_list)
        
        # Agregar BOM UTF-8
        csv_content = '\ufeff' + output.getvalue()
        
        from flask import Response
        return Response(
            csv_content,
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename=proyectos.csv'}
        )
    finally:
        conn.close()


@app.route('/api/exportar-proyecto-documentos-csv/<codigo>')
@login_required
def api_exportar_proyecto_documentos_csv(codigo):
    """Exporta los documentos de un proyecto específico a CSV"""
    import io
    import csv
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Agregar cuenta_corriente del cliente
        cursor.execute('''
            SELECT d.numero_doc, d.tipo_doc, d.fecha_emision, c.razon_social, 
                   d.valor_neto, d.iva, d.valor_total, d.estado, d.forma_pago, c.cuenta_corriente
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE d.proyecto_codigo = ?
            ORDER BY d.fecha_emision DESC
        ''', (codigo,))
        rows = cursor.fetchall()
        
        # Crear CSV con codificación UTF-8 BOM para Excel
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator='\n')
        
        # Obtener nombre del proyecto
        proyecto = cursor.execute('SELECT nombre FROM proyectos WHERE codigo = ?', (codigo,)).fetchone()
        proyecto_nombre = proyecto[0] if proyecto else 'Proyecto'
        
        # Encabezado
        writer.writerow([f'Documentos del Proyecto: {proyecto_nombre}'])
        writer.writerow(['Código de Proyecto', codigo])
        writer.writerow(['Fecha', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow(['Total de documentos', len(rows)])
        writer.writerow([''])
        
        writer.writerow(['N° Doc', 'Tipo', 'Fecha', 'Cliente', 'Valor Neto', 'IVA', 'Total', 'Estado', 'Forma Pago', 'Cuenta Corriente'])
        
        total_neto = 0
        total_iva = 0
        total_general = 0
        
        for row in rows:
            row_list = list(row)
            for i in [4, 5, 6]:
                if row_list[i]:
                    row_list[i] = f"{float(row_list[i]):,.2f}".replace(',', '.')
            writer.writerow(row_list)
            total_neto += float(row[4]) if row[4] else 0
            total_iva += float(row[5]) if row[5] else 0
            total_general += float(row[6]) if row[6] else 0
        
        writer.writerow([''])
        writer.writerow(['TOTALES'])
        writer.writerow(['', '', '', 'Total Neto', f"{total_neto:,.2f}".replace(',', '.')])
        writer.writerow(['', '', '', 'Total IVA', f"{total_iva:,.2f}".replace(',', '.')])
        writer.writerow(['', '', '', 'Total General', f"{total_general:,.2f}".replace(',', '.')])
        
        # Agregar BOM UTF-8
        csv_content = '\ufeff' + output.getvalue()
        
        from flask import Response
        return Response(
            csv_content,
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename=proyecto_{codigo}_documentos.csv'}
        )
    finally:
        conn.close()


# ==================== APIs para Reportes Básicos ====================

@app.route('/api/reporte-resumen')
@login_required
def api_reporte_resumen():
    """Devuelve resumen general de documentos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Total facturado (facturas y boletas)
        cursor.execute('''
            SELECT COALESCE(SUM(valor_total), 0) FROM documentos 
            WHERE tipo_doc IN ('FAC', 'BOL')
        ''')
        total_facturado = cursor.fetchone()[0]
        
        # Total pagado
        cursor.execute('''
            SELECT COALESCE(SUM(valor_total), 0) FROM documentos 
            WHERE tipo_doc IN ('FAC', 'BOL') AND estado = 'Pagado'
        ''')
        total_pagado = cursor.fetchone()[0]
        
        # Total pendiente
        cursor.execute('''
            SELECT COALESCE(SUM(valor_total), 0) FROM documentos 
            WHERE tipo_doc IN ('FAC', 'BOL') AND estado = 'Pendiente'
        ''')
        total_pendiente = cursor.fetchone()[0]
        
        # Total notas de crédito
        cursor.execute('''
            SELECT COALESCE(SUM(valor_total), 0) FROM documentos WHERE tipo_doc = 'NC'
        ''')
        total_nc = cursor.fetchone()[0]
        
        # Total notas de débito
        cursor.execute('''
            SELECT COALESCE(SUM(valor_total), 0) FROM documentos WHERE tipo_doc = 'ND'
        ''')
        total_nd = cursor.fetchone()[0]
        
        # Conteos
        cursor.execute('SELECT COUNT(*) FROM documentos WHERE estado = "Pagado"')
        count_pagados = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM documentos WHERE estado = "Pendiente"')
        count_pendientes = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM documentos WHERE tipo_doc = "NC"')
        count_nc = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM documentos WHERE estado = "Anulado"')
        count_anulados = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM documentos')
        total_documentos = cursor.fetchone()[0]
        
        return jsonify({
            'total_facturado': total_facturado,
            'total_pagado': total_pagado,
            'total_pendiente': total_pendiente,
            'total_nc': total_nc,
            'total_nd': total_nd,
            'count_pagados': count_pagados,
            'count_pendientes': count_pendientes,
            'count_nc': count_nc,
            'count_anulados': count_anulados,
            'total_documentos': total_documentos
        })
    finally:
        conn.close()


@app.route('/api/reporte-nc-nd')
@login_required
def api_reporte_nc_nd():
    """Devuelve lista de notas de crédito y débito"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT d.id, d.numero_doc, d.tipo_doc, d.fecha_emision, d.cliente_rut,
                   d.descripcion, d.valor_neto, d.iva, d.valor_total, d.estado,
                   c.razon_social, c.email as cliente_email
            FROM documentos d
            LEFT JOIN clientes c ON d.cliente_rut = c.rut
            WHERE d.tipo_doc IN ('NC', 'ND')
            ORDER BY d.fecha_emision DESC
        ''')
        rows = cursor.fetchall()
        documentos = []
        for row in rows:
            documentos.append({
                'id': row[0],
                'numero_doc': row[1],
                'tipo_doc': row[2],
                'fecha_emision': row[3],
                'cliente_rut': row[4],
                'descripcion': row[5],
                'valor_neto': row[6],
                'iva': row[7],
                'valor_total': row[8],
                'estado': row[9],
                'razon_social': row[10] or 'Cliente no encontrado',
                'cliente_email': row[11],
                'motivo_nc_nd': row[5],  # Usar descripción como motivo
                'doc_referencia_tipo': None,
                'doc_referencia_numero': None
            })
        return jsonify(documentos)
    finally:
        conn.close()


@app.route('/api/documento-nuevo', methods=['POST'])
@login_required
def api_documento_nuevo():
    """Crear documento con la nueva estructura"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        data = request.get_json()
        
        # Obtener IDs de las tablas de referencia
        tipo_doc_codigo = data.get('tipo_doc', 'FAC')
        cursor.execute('SELECT id, afecto_iva FROM tipo_documento WHERE codigo = ?', (tipo_doc_codigo,))
        tipo_doc = cursor.fetchone()
        tipo_documento_id = tipo_doc[0] if tipo_doc else None
        afecto_iva = tipo_doc[1] if tipo_doc else True
        
        forma_pago_codigo = data.get('forma_pago', 'CONT')
        cursor.execute('SELECT id, dias_plazo FROM forma_pago WHERE codigo = ?', (forma_pago_codigo,))
        forma_pago = cursor.fetchone()
        forma_pago_id = forma_pago[0] if forma_pago else None
        dias_plazo = forma_pago[1] if forma_pago else 0
        
        # Estado pendiente por defecto
        cursor.execute('SELECT id FROM estado_documento WHERE codigo = ?', ('PEN',))
        estado = cursor.fetchone()
        estado_id = estado[0] if estado else None
        
        # Obtener empresa_persona_id desde RUT
        cliente_rut = normalize_rut(data.get('cliente_rut', ''))
        cursor.execute('SELECT id FROM empresa_persona WHERE rut = ?', (cliente_rut,))
        cursor.execute('SELECT id FROM empresa_persona WHERE rut = ?', (cliente_rut,))
        empresa = cursor.fetchone()
        empresa_persona_id = empresa[0] if empresa else None
        
        # Obtener proyecto_id si se proporciona
        proyecto_id = None
        if data.get('proyecto_codigo'):
            cursor.execute('SELECT id FROM proyectos WHERE codigo = ?', (data.get('proyecto_codigo'),))
            proyecto = cursor.fetchone()
            proyecto_id = proyecto[0] if proyecto else None
        
        # Calcular valores
        valor_neto = float(data.get('valor_neto', 0))
        iva = valor_neto * 0.19 if afecto_iva else 0
        valor_total = valor_neto + iva
        
        # Calcular fecha vencimiento
        from datetime import datetime, timedelta
        fecha_emision = data.get('fecha_emision', datetime.now().strftime('%Y-%m-%d'))
        fecha_venc = (datetime.strptime(fecha_emision, '%Y-%m-%d') + timedelta(days=dias_plazo)).strftime('%Y-%m-%d')
        
        # Obtener siguiente número de documento
        cursor.execute('SELECT COALESCE(MAX(numero_doc), 0) + 1 FROM documento WHERE tipo_documento_id = ?', 
                      (tipo_documento_id,))
        numero_doc = data.get('numero_doc') or cursor.fetchone()[0]
        
        # Insertar documento
        cursor.execute('''
            INSERT INTO documento (numero_doc, tipo_documento_id, fecha_emision, fecha_vencimiento,
                                  empresa_persona_id, proyecto_id, forma_pago_id, estado_id,
                                  valor_neto, iva, valor_total, observaciones)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            numero_doc, tipo_documento_id, fecha_emision, fecha_venc,
            empresa_persona_id, proyecto_id, forma_pago_id, estado_id,
            valor_neto, iva, valor_total, data.get('descripcion', '')
        ))
        documento_id = cursor.lastrowid
        
        # Insertar detalles si se proporcionan
        detalles = data.get('detalles', [])
        for i, detalle in enumerate(detalles):
            producto_id = detalle.get('producto_id')
            cursor.execute('''
                INSERT INTO detalle_documento (documento_id, producto_servicio_id, descripcion,
                                              cantidad, precio_unitario, subtotal, orden)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                documento_id, producto_id, detalle.get('descripcion', ''),
                detalle.get('cantidad', 1), detalle.get('precio_unitario', 0),
                detalle.get('cantidad', 1) * detalle.get('precio_unitario', 0), i
            ))
        
        # Mantener compatibilidad con tabla documentos antigua
        cursor.execute('''
            INSERT INTO documentos (numero_doc, tipo_doc, fecha_emision, cliente_rut,
                                   descripcion, valor_neto, iva, valor_total, estado, 
                                   forma_pago, proyecto_codigo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            numero_doc, tipo_doc_codigo, fecha_emision, cliente_rut,
            data.get('descripcion', ''), valor_neto, iva, valor_total,
            'Pendiente', forma_pago_codigo, data.get('proyecto_codigo', '')
        ))
        
        conn.commit()
        return jsonify({
            'success': True,
            'id': documento_id,
            'numero_doc': numero_doc,
            'message': f'Documento {tipo_doc_codigo} #{numero_doc} creado correctamente'
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/documento/<int:doc_id>/detalles')
@login_required
def api_documento_detalles(doc_id):
    """Obtener detalles de un documento"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT dd.id, dd.descripcion, dd.cantidad, dd.precio_unitario, dd.subtotal,
                   ps.codigo as producto_codigo, ps.nombre as producto_nombre
            FROM detalle_documento dd
            LEFT JOIN producto_servicio ps ON dd.producto_servicio_id = ps.id
            WHERE dd.documento_id = ?
            ORDER BY dd.orden
        ''', (doc_id,))
        rows = cursor.fetchall()
        return jsonify([{
            'id': r[0], 'descripcion': r[1], 'cantidad': r[2],
            'precio_unitario': r[3], 'subtotal': r[4],
            'producto_codigo': r[5], 'producto_nombre': r[6]
        } for r in rows])
    finally:
        conn.close()


@app.route('/api/estructura-bd')
@login_required
def api_estructura_bd():
    """Devuelve la estructura de la base de datos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row[0] for row in cursor.fetchall()]
        
        estructura = {}
        for table in tables:
            cursor.execute(f"PRAGMA table_info({table})")
            columns = cursor.fetchall()
            estructura[table] = [{'name': col[1], 'type': col[2], 'pk': bool(col[5])} for col in columns]
        
        return jsonify(estructura)
    finally:
        conn.close()


# ============================================================
# ENDPOINT DE DEBUG
# ============================================================

@app.route('/test-simple')
def test_simple():
    """Página de prueba simple del API"""
    return render_template('test_simple.html')

@app.route('/clientes-dev')
def clientes_dev():
    """Versión de desarrollo sin autenticación"""
    return render_template('clientes_dev.html')


@app.route('/clientes-public')
def clientes_public():
    """Página pública para debug: muestra la plantilla de clientes sin requerir login.
    Útil para verificar que el HTML se renderiza con filas de clientes (activos) en el servidor.
    """
    bancos_chile = [
        'Banco de Chile', 'BCI', 'Scotiabank', 'BBVA', 'Santander', 'Itaú', 'Banco Ripley', 'Banco Falabella'
    ]
    conn = get_db_connection()
    clientes_list = rows_to_dicts(conn.execute("SELECT * FROM clientes WHERE activo = 1 ORDER BY razon_social").fetchall())
    conn.close()
    return render_template('clientes.html', clientes=clientes_list, bancos=sorted(bancos_chile))

@app.route('/test-api')
@login_required
def test_api():
    """Página de prueba del API de clientes"""
    return render_template('test_api.html')

@app.route('/api/debug/clientes-db')
def debug_clientes_db():
    """Endpoint de debug para verificar clientes en la BD"""
    conn = sqlite3.connect(app.config['DATABASE'])
    cursor = conn.cursor()
    try:
        # Contar total de clientes
        total = cursor.execute("SELECT COUNT(*) FROM clientes").fetchone()[0]
        
        # Obtener todos los clientes sin filtro
        rows = cursor.execute(
            "SELECT id, rut, razon_social, email, activo FROM clientes"
        ).fetchall()
        
        clientes = [dict(row) for row in rows]
        
        return jsonify({
            'total': total,
            'clientes': clientes,
            'primera_fila': clientes[0] if clientes else None
        })
    except Exception as e:
        return jsonify({'error': str(e)})
    finally:
        conn.close()


if __name__ == '__main__':
    # En producción (PythonAnywhere) este bloque no se ejecuta
    # PythonAnywhere usa WSGI directamente
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(debug=debug_mode, port=5000)